from __future__ import annotations

import hashlib
import io
import json
import re
import socket
import ssl
import zipfile
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

DATA_HUB = "https://www.aofm.gov.au/data-hub"
PINNED_BONDS_XLSX = "https://www.aofm.gov.au/sites/default/files/2025-06-20/treasury%20bonds%20-%20issuance.xlsx"
TIMEOUT = 8
BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-AU,en;q=0.9",
    "Cache-Control": "no-cache",
}


def emit(obj):
    print(json.dumps(obj, default=str), flush=True)


def transport_probe(host: str = "www.aofm.gov.au") -> dict:
    out = {"host": host}
    try:
        addrs = socket.getaddrinfo(host, 443, type=socket.SOCK_STREAM)
        out["resolved_ips"] = list(dict.fromkeys(a[4][0] for a in addrs))
        raw = socket.create_connection((host, 443), timeout=TIMEOUT)
        ctx = ssl.create_default_context()
        tls = ctx.wrap_socket(raw, server_hostname=host)
        out["tls_ok"] = True
        out["tls_version"] = tls.version()
        out["peer_ip"] = tls.getpeername()[0]
        tls.close()
    except Exception as exc:
        out["tls_ok"] = False
        out["error"] = repr(exc)
    emit({"transport_probe": out})
    return out


def req(method: str, url: str, headers=None, stream=False) -> requests.Response:
    r = requests.request(
        method,
        url,
        timeout=TIMEOUT,
        allow_redirects=True,
        headers=headers or BROWSER_HEADERS,
        stream=stream,
    )
    emit({
        "method": method,
        "request_url": url,
        "status": r.status_code,
        "final_url": r.url,
        "content_type": r.headers.get("content-type"),
        "content_length_header": r.headers.get("content-length"),
        "server": r.headers.get("server"),
    })
    r.raise_for_status()
    return r


def discover_bond_xlsx(html: str) -> list[str]:
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', html, flags=re.I)
    urls = [urljoin(DATA_HUB, h) for h in hrefs]
    return list(dict.fromkeys(
        u for u in urls
        if "treasury" in u.lower()
        and "bond" in u.lower()
        and u.lower().split("?")[0].endswith(".xlsx")
    ))


def validate_xlsx(data: bytes) -> dict:
    if len(data) < 4 or data[:2] != b"PK":
        raise RuntimeError(f"not XLSX/ZIP; first bytes={data[:20]!r}")
    if not zipfile.is_zipfile(io.BytesIO(data)):
        raise RuntimeError("PK present but payload is not a valid ZIP/XLSX")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sample = {}
    for ws in wb.worksheets[:3]:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 15), values_only=True):
            rows.append([v.isoformat() if hasattr(v, "isoformat") else v for v in row[:15]])
        sample[ws.title] = rows
    return {"sheet_names": wb.sheetnames, "sample": sample}


def attempt(label, fn):
    try:
        value = fn()
        emit({"probe": label, "ok": True})
        return True, value, None
    except Exception as exc:
        emit({"probe": label, "ok": False, "error": repr(exc)})
        return False, None, repr(exc)


def main() -> int:
    out = Path("aofm_test_output")
    out.mkdir(exist_ok=True)
    report = {"transport": transport_probe(), "probes": {}}

    ok, hub, err = attempt("browser_get_data_hub", lambda: req("GET", DATA_HUB))
    report["probes"]["browser_get_data_hub"] = {"ok": ok, "error": err}
    discovered = discover_bond_xlsx(hub.text) if ok else []
    if ok:
        report["probes"]["browser_get_data_hub"]["discovered"] = discovered

    ok, head, err = attempt("browser_head_xlsx", lambda: req("HEAD", PINNED_BONDS_XLSX))
    report["probes"]["browser_head_xlsx"] = {"ok": ok, "error": err}

    range_headers = dict(BROWSER_HEADERS)
    range_headers["Range"] = "bytes=0-1023"
    range_headers["Accept"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8"
    ok, ranged, err = attempt(
        "browser_range_xlsx",
        lambda: req("GET", PINNED_BONDS_XLSX, headers=range_headers),
    )
    report["probes"]["browser_range_xlsx"] = {"ok": ok, "error": err}
    if ok:
        report["probes"]["browser_range_xlsx"].update({
            "status": ranged.status_code,
            "bytes": len(ranged.content),
            "first_bytes_hex": ranged.content[:16].hex(),
        })

    xlsx_headers = dict(BROWSER_HEADERS)
    xlsx_headers["Accept"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*;q=0.8"
    ok, full, err = attempt(
        "browser_full_xlsx",
        lambda: req("GET", PINNED_BONDS_XLSX, headers=xlsx_headers),
    )
    report["probes"]["browser_full_xlsx"] = {"ok": ok, "error": err}
    if ok:
        meta = validate_xlsx(full.content)
        sha = hashlib.sha256(full.content).hexdigest()
        (out / "treasury_bonds_issuance.xlsx").write_bytes(full.content)
        report["probes"]["browser_full_xlsx"].update({
            "status": full.status_code,
            "final_url": full.url,
            "content_type": full.headers.get("content-type"),
            "byte_length": len(full.content),
            "sha256": sha,
            **meta,
        })

    report["success"] = bool(report["probes"]["browser_full_xlsx"]["ok"])
    (out / "result.json").write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, indent=2, default=str), flush=True)
    return 0 if report["success"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
